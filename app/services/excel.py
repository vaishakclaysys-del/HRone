from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.models import Candidate, CandidateScreening, InterviewScore, CandidateProgress, Interview, User


# ---------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------

def _normalize_excel_header(value: object) -> str:
    return "".join(
        ch for ch in str(value or "").strip().lower()
        if ch.isalnum()
    )


def _clean_cell(value: object) -> str:
    return str(value).strip() if value is not None else ""


# ---------------------------------------------------------
# Build screening Excel bytes
# ---------------------------------------------------------

def build_screening_excel_bytes(db: Session) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Screening Results"

    ws.append([
        "ID", "AIML Relevant", "ML Foundations", "Practical ML",
        "LLM Stack", "Implementation", "Problem Solving",
        "Attitude", "Total Score", "Recommendation", "Summary",
    ])

    rows = db.execute(
        select(Candidate, CandidateScreening)
        .join(CandidateScreening, CandidateScreening.candidate_id == Candidate.id)
        .order_by(Candidate.id.asc())
    ).all()

    for candidate, screening in rows:
        total = float(screening.fit_score or 0.0)
        ws.append([
            candidate.id,
            "Yes" if total >= 70 else "No",
            0, 0, 0, 0, 0, 0,
            total,
            screening.fit_tag,
            screening.recommended_problem,
        ])

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ---------------------------------------------------------
# Parse HR master Excel rows
# ---------------------------------------------------------

def parse_hr_master_rows(
    hr_file_bytes: bytes,
) -> tuple[list[dict[str, str]], int]:
    """Parse standard HR master rows from uploaded Excel."""

    workbook    = load_workbook(filename=BytesIO(hr_file_bytes), data_only=True)
    parsed_rows: list[dict[str, str]] = []
    skipped_rows = 0

    for sheet in workbook.worksheets:
        rows = sheet.iter_rows(values_only=True)

        try:
            header_row = next(rows)
        except StopIteration:
            continue

        headers   = [_normalize_excel_header(cell) for cell in header_row]
        id_idx    = next((i for i, h in enumerate(headers) if h == "id"), None)
        name_idx  = next((i for i, h in enumerate(headers) if h == "name"), None)
        phone_idx = next((i for i, h in enumerate(headers) if h in {"mobile", "phone", "contact", "mobilenumber"}), None)
        email_idx = next((i for i, h in enumerate(headers) if h == "email"), None)

        if id_idx is None or name_idx is None or phone_idx is None:
            continue

        for row in rows:
            if not row:
                continue
            if id_idx >= len(row) or phone_idx >= len(row) or name_idx >= len(row):
                skipped_rows += 1
                continue

            raw_id    = _clean_cell(row[id_idx])
            raw_name  = _clean_cell(row[name_idx])
            raw_phone = _clean_cell(row[phone_idx])
            raw_email = _clean_cell(row[email_idx]) if email_idx is not None and email_idx < len(row) else ""

            if not raw_id or not raw_name or not raw_phone:
                skipped_rows += 1
                continue
            if not raw_id.isdigit():
                skipped_rows += 1
                continue

            parsed_rows.append({
                "id":    raw_id,
                "name":  raw_name,
                "phone": raw_phone,
                "email": raw_email,
            })

    return parsed_rows, skipped_rows


# ---------------------------------------------------------
# Sync candidates from HR rows
# ---------------------------------------------------------

def sync_candidates_from_hr_rows(
    db: Session,
    rows: list[dict[str, str]],
) -> tuple[int, int, int]:
    """
    Upsert candidates by HR ID.
    ID is the sole identity key; phone is stored for reference only.
    Returns (created, updated, skipped).
    """

    created = updated = skipped = 0

    for row in rows:
        hr_id = int(row["id"])
        candidate = db.get(Candidate, hr_id)

        if candidate is None:
            db.add(Candidate(
                id     = hr_id,
                name   = row["name"],
                phone  = row["phone"] or None,
                email  = row["email"] or None,
                skills = "",
                stage  = 1,
                status = "new",
            ))
            created += 1
        else:
            candidate.name  = row["name"]
            candidate.phone = row["phone"] or None
            candidate.email = row["email"] or None
            updated += 1

    db.commit()
    return created, updated, skipped

def build_candidates_export_excel_bytes(db: Session) -> bytes:
    """
    Build the 'final_candidates.xlsx' export workbook for candidates
    who have reached the interview stage (passed, failed, or
    scheduled), including their progress timeline, interviewers,
    interview dates, scores, and recommendations.
    """
    candidates = list(
        db.scalars(
            select(Candidate).where(
                Candidate.status.in_(
                    [
                        "interview_passed",
                        "interview_failed",
                        "interview_scheduled",
                    ]
                )
            )
        )
    )
 
    wb = Workbook()
    ws = wb.active
    ws.title = "Candidates"
 
    ws.append(
        [
            "Candidate Name",
            "Email",
            "Phone",
            "Skills",
            "Stage",
            "Status",
            "Timeline",
            "Progress Entries",
            "Interviewers",
            "Interview Dates",
            "Scores",
            "Recommendations",
        ]
    )
 
    for candidate in candidates:
 
        progress_entries = list(
            db.scalars(
                select(CandidateProgress)
                .where(CandidateProgress.candidate_id == candidate.id)
                .order_by(CandidateProgress.created_at)
            )
        )
 
        interviews = list(
            db.scalars(
                select(Interview)
                .where(Interview.candidate_id == candidate.id)
                .order_by(Interview.created_at.desc())
            )
        )
 
        interviewer_names = []
        interview_dates = []
        scores = []
        recommendations = []
 
        for interview in interviews:
 
            interviewer = db.get(User, interview.interviewer_id)
 
            if interviewer:
                interviewer_names.append(interviewer.full_name)
 
            if interview.scheduled_at:
                interview_dates.append(str(interview.scheduled_at))
 
            interview_scores = db.scalars(
                select(InterviewScore).where(
                    InterviewScore.interview_id == interview.id
                )
            )
 
            for score in interview_scores:
                scores.append(str(score.score))
                recommendations.append(score.recommendation or "")
 
        ws.append(
            [
                candidate.name,
                candidate.email,
                candidate.phone,
                candidate.skills,
                candidate.stage,
                candidate.status,
                "\n".join(
                    f"{p.created_at.strftime('%Y-%m-%d %H:%M')} - {p.event}"
                    for p in progress_entries
                ),
                "\n".join(
                    f"{p.event} - {p.details or ''}"
                    for p in progress_entries
                ),
                ", ".join(interviewer_names),
                ", ".join(interview_dates),
                ", ".join(scores),
                ", ".join(recommendations),
            ]
        )
 
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
 
    return stream.getvalue()
 